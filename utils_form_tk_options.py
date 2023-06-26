import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl import drawing

import matplotlib.pyplot as plt
# import seaborn as sns
# %matplotlib inline
from matplotlib.colors import ListedColormap, BoundaryNorm

from utils_form_tk_enrichment_options import preprocess_tkbd_options
from utils_io import save_df_lst_to_excel, rename_sheet, form_str_date
from utils_io import logger

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    from utils_io import logger

from  local_dictionaries import sevice_sections, service_chapters, service_types_A, service_types_B, service_classes_A, service_classes_B
from  local_dictionaries import dict_ath_anatomy, dict_ath_therapy, dict_ath_pharm, dict_ath_chemical

def read_enriched_tk_data(path_tkbd_processed, fn_tk_bd):
    df_services = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'Услуги')
    print(df_services.shape)
    display(df_services.head(2))
    df_LP = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'ЛП')
    print(df_desc.shape)
    display(df_LP.head(2))
    df_RM = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'РМ')
    print(df_RM.shape)
    display(df_RM.head(2))
    return df_services, df_LP, df_RM




def create_tk_models_dict_02(model, xls_file, tk_code=7777777, tk_name='tk_test', profile='profile_test', tk_models = {} ):
    # tk_models = {}
    max_len = 40

    if tk_models.get(tk_name) is None:
        tk_models[tk_name] = {}
    tk_models[tk_name]['Код ТК'] = tk_code
    tk_models[tk_name]['Профиль'] = profile
    tk_models[tk_name]['Наименование ТК (короткое)'] = tk_name[:max_len]
    # 'Модели': [{'Модель пациента': 'Техно',
    #             'Файл Excel': 'Аденома_предстательной_железы_Склероз_шейки_мочевого_пузыря_Стриктура '
    #                           'техно.xlsx'},
    #           {'Модель пациента': 'База',
    #             'Файл Excel': 'Аденома_предстательной_железы_Склероз_шейки_мочевого_пузыря_Стриктура.xlsx'}],

    # tk_models[tk_name]['Модели'] = [models]
        # tk_models.setdefault('tk_name', []).append(row['Наименование ТК'])
    #tk_models[tk_name]['Модели'].append (dict(zip(['Модель пациента', 'Файл Excel',
    #  '#Название листа в файле Excel', 'Услуги', 'ЛП', 'РМ'], row.values[4:])))
    # tk_models[tk_name]['Модели'].append (dict(zip(['Модель пациента', 'Файл Excel',], row.values[4:6])))
    tk_models[tk_name]['Модели'] = []
    for i_m, model in enumerate(models):
        tk_models[tk_name]['Модели'].append (dict(zip(['Модель пациента', 'Файл Excel',], [model, xls_files[i_m]])))

    return tk_models

def format_excel_cols_short(ws, format_cols, auto_filter=False):
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    last_cell = ws.cell(row=1, column=len(format_cols))
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    if auto_filter:
        ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(format_cols):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    return ws


def change_order_base_techno(new_columns_02):
    if 'База' in new_columns_02 and 'Техно' in new_columns_02:
        i_base = new_columns_02.index('База')
        i_techno = new_columns_02.index('Техно')
        # print(i_base, i_techno)
        if i_techno < i_base:
            new_columns_03 = [col for col in new_columns_02 if col not in ['Техно', 'База']]
            # print(new_columns_03)
            if i_base > 0: i_base -= 1
            new_columns_03.insert(i_base, 'Техно')
            new_columns_03.insert(i_base, 'База')


            return new_columns_03

        else: return new_columns_02
    else: return new_columns_02

def reorder_columns_by_models(new_columns_02, model_names):
    if model_names[0] in new_columns_02 and model_names[1] in new_columns_02:
        i_first = new_columns_02.index(model_names[0])
        i_second = new_columns_02.index(model_names[1])
        # print(i_base, i_techno)
        if i_second < i_first:
            new_columns_03 = [col for col in new_columns_02 if col not in model_names]
            # print(new_columns_03)
            if i_first > 0: i_first -= 1
            new_columns_03.insert(i_first, model_names[1])
            new_columns_03.insert(i_first, model_names[0])


            return new_columns_03

        else: return new_columns_02
    else: return new_columns_02

def extract_names_from_code_service(code, debug=False):
    section_name, type_name, class_name = None, None, None
    if (type(code)!= str) or ((type(code)==str) and (len(code)==0)): return section_name, type_name, class_name
    section_name = sevice_sections.get(code[0])
    if len(code)>=3:
        if code[0] == 'A':
            if debug: print('A')
            type_name = service_types_A.get(code[1:3])
        else: type_name = service_types_B.get(code[1:3])
        if len(code)>=6:
            if code[0] == 'A':
                class_name = service_classes_A.get(code[4:6])
            else: class_name = service_classes_B.get(code[4:7])
        else: return section_name, type_name, class_name
    else: return section_name, type_name, class_name

    return section_name, type_name, class_name




def extract_name_groups_ATH(s, debug = False):
    ath_anatomy_code, ath_anatomy, ath_therapy_code, ath_therapy, ath_pharm_code, ath_pharm, ath_chemical_code, ath_chemical = \
        None, None, None, None, None, None, None, None
    if type(s) is None or ((type(s)==float) and np.isnan(s)) or (type(s)!=str)  or ((type(s)==str) and (len(s)==0)):
        return None, None, None, None, None, None, None, None
    ath_anatomy_code = s[0]
    ath_anatomy = dict_ath_anatomy.get(ath_anatomy_code)
    if len(s)>=3:
        ath_therapy_code = s[0:3]
        ath_therapy = dict_ath_therapy.get(ath_therapy_code)
        if len(s)>=4:
            ath_pharm_code = s[0:4]
            ath_pharm = dict_ath_pharm.get(ath_pharm_code)
            if len(s)>=5:
                ath_chemical_code = s[0:5]
                ath_chemical = dict_ath_chemical.get(ath_chemical_code)
    return ath_anatomy, ath_therapy, ath_pharm, ath_chemical




# import pandas as pd
from utils_form_tk_enrichment_options import preprocess_tkbd_options
from utils_io import form_str_date, format_excel_sheet_cols, add_sheet_to_excel_from_df, save_df_lst_to_excel

def group_services(df_services, freq_threshold, group_code_col ='Код типа', group_name_col='Тип'):

    code_col = 'Код услуги по Номенклатуре медицинских услуг (Приказ МЗ № 804н)'
    name_col = 'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)'
    freq_col = 'Усредненная частота предоставления'
    multi_col = 'Усредненная кратность применения'
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel']
    main_cols = ['Код услуги по Номенклатуре медицинских услуг (Приказ МЗ № 804н)',
       'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)',
       'Усредненная частота предоставления',
       'Усредненная кратность применения'] #, 'Код раздела', 'Раздел', 'Код типа'

    df_g_services1 = df_services[df_services[freq_col]>=freq_threshold]
    df_g_services2 = df_services[df_services[freq_col]<freq_threshold]
    print(df_services.shape, df_g_services1.shape, df_g_services2.shape)
    total_source_positions = df_services.shape[0]
    total_positions_ge = df_g_services1.shape[0]   # (greater|equal)
    total_positions_less = df_g_services2.shape[0]
    logger.info(f"Исходное количество позиций: {total_source_positions}")
    logger.info(f"Позиций >= порога: {total_positions_ge}")
    logger.info(f"Позиций < порога: {total_positions_less}")

    # display(df_g_services2.groupby(group_col).sum(freq_col))
    # display(df_g_services2[list(df_g_services1.columns)].groupby(group_col).sum(freq_col).reset_index().head())
    groupby_cols = head_cols + [group_code_col, group_name_col]
    df_g_services2_g = df_g_services2[ groupby_cols + [freq_col, multi_col]
                                      ].groupby(groupby_cols).agg({freq_col: 'sum', multi_col: 'mean'}).reset_index()
    # print("df_g_services2_g.columns:", df_g_services2_g.columns)
    # display(df_g_services2_g.head(2))
    df_g_services2_g[code_col] = df_g_services2_g[group_code_col].apply(lambda x: x + ('.AA' if x.startswith('A') else '.BBB'))
    df_g_services2_g[name_col] = df_g_services2_g[group_name_col].apply(lambda x: x.capitalize() + '.*')

    # print("df_g_services1.columns:", df_g_services1.columns)
    # print(df_g_services2_g.shape)
    # display(df_g_services2_g.head(2))
    df_g_services2_g.drop(columns= [group_code_col, group_name_col], inplace=True)
    df_g_services = pd.concat([df_g_services1[list(df_g_services2_g.columns)], df_g_services2_g])

    df_g_services = df_g_services[head_cols + [code_col, name_col, freq_col, multi_col]]
    df_g_services = df_g_services.sort_values(by=[code_col], ascending=True)
    total_proc_positions = df_g_services.shape[0]
    logger.info(f"Итоговое количество позиций: {total_proc_positions}")
    print(df_g_services.shape)

    return df_g_services, total_source_positions, total_positions_ge, total_positions_less, total_proc_positions

def group_services_02(df_services, freq_threshold, group_code_col ='Код типа', group_name_col='Тип'):

    code_col = 'Код услуги по Номенклатуре медицинских услуг (Приказ МЗ № 804н)'
    name_col = 'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)'
    freq_col = 'Усредненная частота предоставления'
    multi_col = 'Усредненная кратность применения'
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel']
    main_cols = ['Код услуги по Номенклатуре медицинских услуг (Приказ МЗ № 804н)',
       'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)',
       freq_col, multi_col, ] #, 'Код раздела', 'Раздел', 'Код типа'

    df_g_services1 = df_services[df_services[freq_col]>=freq_threshold]
    df_g_services2 = df_services[df_services[freq_col]<freq_threshold]
    print(df_services.shape, df_g_services1.shape, df_g_services2.shape)
    total_source_positions = df_services.shape[0]
    total_positions_ge = df_g_services1.shape[0]   # (greater|equal)
    total_positions_less = df_g_services2.shape[0]
    logger.info(f"Исходное количество позиций: {total_source_positions}")
    logger.info(f"Позиций >= порога: {total_positions_ge}")
    logger.info(f"Позиций < порога: {total_positions_less}")

    # display(df_g_services2.groupby(group_col).sum(freq_col))
    # display(df_g_services2[list(df_g_services1.columns)].groupby(group_col).sum(freq_col).reset_index().head())
    groupby_cols = head_cols + [group_code_col, group_name_col]
    # df_g_services2_g = df_g_services2[ groupby_cols + [freq_col, multi_col]
    #                                   ].groupby(groupby_cols).agg({freq_col: 'sum', multi_col: 'mean'}).reset_index()
    df_g_services2_g = df_g_services2[ groupby_cols + [freq_col, multi_col]].groupby(groupby_cols)
    # df_g_services2_g_ri= df_g_services2_g.reset_index()
    lst_by_group = []
    head_values = list(df_g_services2[head_cols].values[0])
    # print("head_values:", head_values)
    for ig, (group_name, group_df) in enumerate(df_g_services2_g):

        # print(group_name)
        # display(group_df.head(2))
        freq_sum = group_df.groupby(groupby_cols)[[freq_col]].sum(freq_col).values[0,0]
        # print("freq_sum:", freq_sum)
        if freq_sum!=0:
            group_df['freq_weight'] = group_df[freq_col]/freq_sum
        else: group_df['freq_weight'] = 0
        group_df['multi_weight'] = group_df['freq_weight']*group_df[multi_col]
        multi_avg = group_df['multi_weight'].sum()
        # print(f"freq_sum: {freq_sum}, multi_avg: {multi_avg}")

        # if ig>1: sys.exit(2)
        code_g = group_df[group_code_col].values[0]
        code_g = code_g + ('.AA' if code_g.startswith('A') else '.BBB')
        name_g = group_df[group_name_col].values[0].capitalize() + '.*'
        # print(f"code_g: {code_g}, name_g: {name_g}")
        lst_by_group.append(head_values + [code_g, name_g, freq_sum, multi_avg])

    # # print("df_g_services1.columns:", df_g_services1.columns)
    # # print(df_g_services2_g.shape)
    # # display(df_g_services2_g.head(2))
    # df_g_services2_g.drop(columns= [group_code_col, group_name_col], inplace=True)
    df_g_services2_g = pd.DataFrame(lst_by_group, columns = head_cols + [code_col, name_col, freq_col, multi_col])
    # display(df_g_services2_g.head(1))
    df_g_services = pd.concat([df_g_services1[list(df_g_services2_g.columns)], df_g_services2_g])

    df_g_services = df_g_services[head_cols + [code_col, name_col, freq_col, multi_col]]
    df_g_services = df_g_services.sort_values(by=[code_col], ascending=True)
    total_proc_positions = df_g_services.shape[0]
    logger.info(f"Итоговое количество позиций: {total_proc_positions}")
    # print(df_g_services.shape)

    return df_g_services, total_source_positions, total_positions_ge, total_positions_less, total_proc_positions

def group_items(df_services, df_LP, df_RM,
        freq_threshold,
):
    df_g_services, df_g_LP, df_g_RM = None, None, None
    dict_stat= {}
    if df_services is not None:
        df_g_services, total_source_positions, total_positions_ge, total_positions_less, total_proc_positions = group_services_02(df_services, freq_threshold)
        dict_stat['Услуги'] = {}
        dict_stat['Услуги']['1. Порог частоты'] = freq_threshold
        dict_stat['Услуги']['2. Исходных позиций'] = total_source_positions
        dict_stat['Услуги']['3. Позиций >= порога'] = total_positions_ge
        dict_stat['Услуги']['4. Позиций < порога'] = total_positions_less
        dict_stat['Услуги']['5. Итоговых позиций'] = total_proc_positions
    return df_g_services, df_g_LP, df_g_RM, dict_stat

def form_tk_options( data_source_dir, data_processed_dir, supp_dict_dir,
                        fn_check_file1, cmp_cols_file_01,
                        fn_smnn_pickle, cmp_sections,
                        profile='profile_test', tk_code=7777777, tk_name='tk_test',
                        models = ['Факт', ],
                        freq_threshold=0.05,

                        ):

    if fn_check_file1 is None:
        logger.error(f"Выберите название файла: в параметрах запуска программы")
        sys.exit(2)
    # df_services, df_LP, df_RM = preprocess_tkbd_options(data_source_dir, fn_tk_bd, data_processed_dir, supp_dict_dir, fn_smnn_pickle)
    df_services, df_LP, df_RM = preprocess_tkbd_options(
                data_source_dir, data_processed_dir, supp_dict_dir,
                fn_check_file1,
                cmp_cols_file_01,

                fn_smnn_pickle, cmp_sections,
                profile, tk_code, tk_name,
                models,
                save_enriched=False,
                )
    if df_services is not None: display(df_services.head(2))
    if df_LP is not None: display(df_LP.head(2))
    if df_RM is not None: display(df_RM.head(2))

    df_g_services, df_g_LP, df_g_RM, dict_stat = group_items(df_services, df_LP, df_RM, freq_threshold)
    df_g_lst, sheet_names_lst, sheet_names_g_lst = [], [], []
    if df_g_services is not None:
        df_g_lst.append(df_g_services)
        sheet_names_lst.append(f"Услуги")
        sheet_names_g_lst.append(f"Услуги_грп_{str(freq_threshold)}")
    if df_g_LP is not None:
        df_g_lst.append(df_g_LP)
        sheet_names_lst.append(f"ЛП")
        sheet_names_g_lst.append(f"ЛП_грп_{str(freq_threshold)}")
    if df_g_RM is not None:
        df_g_lst.append(df_g_RM)
        sheet_names_lst.append(f"РМ")
        sheet_names_g_lst.append(f"РМ_грп_{str(freq_threshold)}")
    str_date = form_str_date ()
    fn_tk = 'groupped_tk_' + str_date + '.xlsx'
    fn_tk = f"groupped_tk_{profile}_{tk_code}.xlsx"

    # fn_save = save_df_lst_to_excel(df_g_lst, sheet_names_g_lst, data_processed_dir, fn_tk)
    # fn_save = save_df_lst_to_excel(df_g_lst, sheet_names_lst, data_processed_dir, fn_tk)
    fn_save = f"groupped_tk_{profile}_{tk_code}_{form_str_date()}.xlsx"
    sections = ['Услуги', 'ЛП', 'РМ']
    format_cols = [
        [30,15,30, 30,30, 15,30, 20,20],
        [],
        [],
    ]

    wb = openpyxl.load_workbook(os.path.join(data_source_dir, fn_check_file1))
    wb.save(os.path.join(data_processed_dir, fn_save))
    for section in sections:
        rename_sheet(data_processed_dir, fn_save, section, section + '_дет')

    #for i_sh, sheet_name in enumerate(sheet_names_g_lst):
    for i_sh, sheet_name in enumerate(sheet_names_lst):
        #if i_sh==0:
        #    add_sheet_to_excel_from_df(df_g_lst[i_sh], sheet_name,  data_source_dir, fn_check_file1, data_processed_dir, fn_save, index=False)
        # else:
        add_sheet_to_excel_from_df(df_g_lst[i_sh], sheet_name,  data_processed_dir, fn_save, data_processed_dir, fn_save, index=False)
        format_excel_sheet_cols(data_processed_dir, fn_save, format_cols[sections.index(sheet_names_lst[i_sh])], sheet_name=sheet_name)
    df_dict_stat = pd.DataFrame(dict_stat).reset_index().rename(columns={'index' : 'Показатели'})
    display(df_dict_stat)
    add_sheet_to_excel_from_df(df_dict_stat, 'Stat',  data_processed_dir, fn_save, data_processed_dir, fn_save, index=False)
    format_excel_sheet_cols(data_processed_dir, fn_save, [30, 10, 10, 10], sheet_name='Stat')

    logger.info(f"Файл '{fn_save}' сохранен в директорию '{data_processed_dir}'")
    return df_g_services, df_g_LP, df_g_RM
