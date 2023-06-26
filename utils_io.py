import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
import requests
from urllib.parse import urlencode
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl import drawing

from matplotlib.colors import ListedColormap, BoundaryNorm


class Logger():
    def __init__(self, name = 'Fuzzy Lookup',
                 strfmt = '[%(asctime)s] [%(levelname)s] > %(message)s', # strfmt = '[%(asctime)s] [%(name)s] [%(levelname)s] > %(message)s'
                 level = logging.INFO,
                 datefmt = '%H:%M:%S', # '%Y-%m-%d %H:%M:%S'
                #  datefmt = '%H:%M:%S %p %Z',

                 ):
        self.name = name
        self.strfmt = strfmt
        self.level = level
        self.datefmt = datefmt
        self.logger = logging.getLogger(name)
        self.logger.setLevel(self.level) #logging.INFO)
        self.offset = datetime.timezone(datetime.timedelta(hours=3))
        # create console handler and set level to debug
        self.ch = logging.StreamHandler()
        self.ch.setLevel(self.level)
        # create formatter
        self.strfmt = strfmt # '[%(asctime)s] [%(levelname)s] > %(message)s'
        self.datefmt = datefmt # '%H:%M:%S'
        # СЃРѕР·РґР°РµРј С„РѕСЂРјР°С‚С‚РµСЂ
        self.formatter = logging.Formatter(fmt=strfmt, datefmt=datefmt)
        self.formatter.converter = lambda *args: datetime.datetime.now(self.offset).timetuple()
        self.ch.setFormatter(self.formatter)
        # add ch to logger
        self.logger.addHandler(self.ch)
logger = Logger().logger
logger.propagate = False

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    # del logger
    logger = Logger().logger
    logger.propagate = False


def unzip_file(path_source, fn_zip, work_path):
    logger.info('Unzip ' + fn_zip + ' start...')

    try:
        with zipfile.ZipFile(path_source + fn_zip, 'r') as zip_ref:
            fn_list = zip_ref.namelist()
            zip_ref.extractall(work_path)
        logger.info('Unzip ' + fn_zip + ' done!')
        return fn_list[0]
    except Exception as err:
        logger.error('Unzip error: ' + str(err))
        sys.exit(2)

def save_df_to_excel(df, path_to_save, fn_main, columns = None, b=0, e=None, index=False, sheet_name=None):
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    logger.info(fn + ' save - start ...')
    if e is None or (e <0):
        e = df.shape[0]
    if columns is None:
        if sheet_name is None:
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index)
        else:
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, sheet_name=sheet_name)
    else:
        if sheet_name is None:
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, columns = columns)
        else:
            df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, columns = columns, sheet_name=sheet_name)
    logger.info(fn + ' saved to ' + path_to_save)
    hfs = get_humanize_filesize(path_to_save, fn)
    logger.info("Size: " + str(hfs))
    return fn

def save_df_lst_to_excel(df_lst, sheet_names_lst, save_path, fn):
    # fn = model + '.xlsx'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_date = fn.replace('.xlsx','')  + '_' + str_date + '.xlsx'

    # with pd.ExcelWriter(os.path.join(path_tkbd_processed, fn_date )) as writer:
    with pd.ExcelWriter(os.path.join(save_path, fn_date )) as writer:

        for i, df in enumerate(df_lst):
            df.to_excel(writer, sheet_name = sheet_names_lst[i], index=False)
    return fn_date



def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size

def restore_df_from_pickle(path_files, fn_pickle):

    if fn_pickle is None:
        logger.error('Restore pickle from ' + path_files + ' failed!')
        sys.exit(2)
    if os.path.exists(os.path.join(path_files, fn_pickle)):
        df = pd.read_pickle(os.path.join(path_files, fn_pickle))
        # logger.info('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' done!')
        logger.info('Restore ' + fn_pickle + ' done!')
        logger.info('Shape: ' + str(df.shape))
    else:
        # logger.error('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' from ' + path_files + ' failed!')
        logger.error('Restore ' + fn_pickle + ' from ' + path_files + ' failed!')
    return df

def get_cols_width_exists(ws):
    cols_width_exists = []
    ws.sheet_state, ws.max_row, ws.max_column
    for ic in range(ws.max_column):
        cell = ws.cell(row=1, column=ic+1)
        cols_width_exists.append(ws.column_dimensions[cell.column_letter].width)
    return cols_width_exists

def format_excel_cols_short(ws, format_cols, auto_filter=False):
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    last_cell = ws.cell(row=1, column=len(format_cols))
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    if auto_filter:
        ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(format_cols):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    return ws

def format_excel_sheet_cols(data_processed_dir, fn_xls, col_width_lst, sheet_name):
    wb = load_workbook(os.path.join(data_processed_dir, fn_xls))
    # ws = wb.active
    ws = wb[sheet_name]
    # l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    border = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
     )


    # ws.filterMode = True
    last_cell = ws.cell(row=1, column=len(col_width_lst))
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(col_width_lst):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    # ft = cell.font
    # ft = Font(bold=False)
    # for row in ws[full_range]: #[1:]
    #     for cell in row:
    #         cell.font = ft
    #         cell.alignment = l_alignment
    #         cell.border = border
    wb.save(os.path.join(data_processed_dir, fn_xls))
def form_str_date():
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    return str_date

def add_sheet_to_excel_from_df(
    df, sheet_name,
    path_source, fn_source,
    path_save, fn_save,
    index=False,
    ):

    wb = load_workbook(os.path.join(path_source, fn_source))
    wb.save(os.path.join(path_save, fn_save))

    with pd.ExcelWriter(os.path.join(path_save, fn_save), mode='a', if_sheet_exists='new') as writer: #  engine='openpyxl',
        # Engine to use for writing. If None, defaults to io.excel.<extension>.writer. NOTE: can only be passed as a keyword argument.
        # Deprecated since version 1.2.0: As the xlwt package is no longer maintained, the xlwt engine will be removed in a future version of pandas.
        # if_sheet_exists{вЂerrorвЂ™, вЂnewвЂ™, вЂreplaceвЂ™, вЂoverlayвЂ™}, default вЂerrorвЂ™
        df.to_excel(writer, sheet_name=sheet_name, index=index)

def rewrite_excel_by_df(
    df_test_serv,
    data_source_dir, data_processed_dir,
    fn_check_file, sheet_name,
    max_sim_entries,
      ):
    wb = load_workbook(os.path.join(data_source_dir, fn_check_file))
    ws = wb[sheet_name]
    cols_width_exists  = get_cols_width_exists(ws)
    # print(cols_width_exists)
    cols_width_new = cols_width_exists
    for _ in range(max_sim_entries):
        cols_width_new += [10., 15., 40.]
    # print(cols_width_new)

    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_save = f"{fn_check_file[:fn_check_file.rfind('.')]}_{str_date}.{fn_check_file.split('.')[-1]}"
    # print(fn_save)
    wb.save(os.path.join(data_processed_dir, fn_save))

    with pd.ExcelWriter(os.path.join(data_processed_dir, fn_save), mode='a', if_sheet_exists='new') as writer: #  engine='openpyxl',
        # Engine to use for writing. If None, defaults to io.excel.<extension>.writer. NOTE: can only be passed as a keyword argument.
        # Deprecated since version 1.2.0: As the xlwt package is no longer maintained, the xlwt engine will be removed in a future version of pandas.
        # if_sheet_exists{вЂerrorвЂ™, вЂnewвЂ™, вЂreplaceвЂ™, вЂoverlayвЂ™}, default вЂerrorвЂ™
        df_test_serv.to_excel(writer, sheet_name=f"{sheet_name}_STS", index=False)
        # СЂР°Р±РѕС‚Р°РµС‚ СЃ Pandas 1.4.1 fail pandas 1.4.4

    format_excel_sheet_cols(data_processed_dir, fn_save, cols_width_new, f"{sheet_name}_STS")

    return fn_save

def rename_sheet(path, fn, sheet_old, sheet_new):
    try:
        wb=openpyxl.load_workbook(os.path.join(path, fn))
        try:
            sheet = wb[sheet_old]
            sheet.title = sheet_new
            wb.save(os.path.join(path, fn))
        except Exception as err:
            print(f"В файле нет листа: '{sheet_old}'")
    except Exception as err:
            print(f"{str(err)}")
